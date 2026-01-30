"""
Excel Sheet Splitter - Backend Tests
테스트 실행: python -m pytest test_app.py -v
"""

import pytest
import os
import sys
import io
import tempfile
import json
from pathlib import Path

# 현재 디렉토리를 Python 경로에 추가
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import app, sanitize_filename, handle_duplicate_filename
import openpyxl

# ==================== FIXTURES ====================

@pytest.fixture
def client():
    """Flask 테스트 클라이언트"""
    app.config['TESTING'] = True
    with app.test_client() as client:
        yield client


@pytest.fixture
def sample_excel_2sheets():
    """2개 시트가 있는 샘플 엑셀 파일"""
    wb = openpyxl.Workbook()
    
    # Sheet1 (기본)
    ws1 = wb.active
    ws1.title = "Sales"
    ws1['A1'] = "Product"
    ws1['B1'] = "Amount"
    ws1['A2'] = "Laptop"
    ws1['B2'] = 1000
    ws1['A3'] = "Mouse"
    ws1['B3'] = 50
    
    # Sheet2
    ws2 = wb.create_sheet("Expenses")
    ws2['A1'] = "Category"
    ws2['B1'] = "Cost"
    ws2['A2'] = "Office Rent"
    ws2['B2'] = 5000
    ws2['A3'] = "Utilities"
    ws2['B3'] = 1000
    
    # 임시 파일로 저장
    temp_dir = tempfile.mkdtemp()
    temp_file = os.path.join(temp_dir, "sample_2sheets.xlsx")
    wb.save(temp_file)
    wb.close()
    
    yield temp_file
    
    # 정리
    import shutil
    shutil.rmtree(temp_dir, ignore_errors=True)


@pytest.fixture
def sample_excel_korean():
    """한글 및 특수문자 시트명이 있는 엑셀 파일"""
    wb = openpyxl.Workbook()
    
    # Sheet1: 한글명
    ws1 = wb.active
    ws1.title = "2024년 매출"
    ws1['A1'] = "날짜"
    ws1['A2'] = "2024-01-01"
    
    # Sheet2: 특수문자
    ws2 = wb.create_sheet("Data !@#$")
    ws2['A1'] = "Info"
    ws2['A2'] = "Test"
    
    # 임시 파일로 저장
    temp_dir = tempfile.mkdtemp()
    temp_file = os.path.join(temp_dir, "sample_korean.xlsx")
    wb.save(temp_file)
    wb.close()
    
    yield temp_file
    
    # 정리
    import shutil
    shutil.rmtree(temp_dir, ignore_errors=True)


# ==================== TEST: UTILITY FUNCTIONS ====================

class TestUtilityFunctions:
    """유틸 함수 테스트"""
    
    def test_sanitize_filename_basic(self):
        """기본 파일명 sanitize"""
        assert sanitize_filename("test.xlsx") == "test.xlsx"
        assert sanitize_filename("my_file.xlsx") == "my_file.xlsx"
    
    def test_sanitize_filename_windows_forbidden(self):
        """Windows 금지 문자 제거"""
        assert "<" not in sanitize_filename("file<name>.xlsx")
        assert ">" not in sanitize_filename("file>name>.xlsx")
        assert ":" not in sanitize_filename("file:name.xlsx")
        assert '"' not in sanitize_filename('file"name.xlsx')
        assert "/" not in sanitize_filename("file/name.xlsx")
        assert "\\" not in sanitize_filename("file\\name.xlsx")
        assert "|" not in sanitize_filename("file|name.xlsx")
        assert "?" not in sanitize_filename("file?name.xlsx")
        assert "*" not in sanitize_filename("file*name.xlsx")
    
    def test_sanitize_filename_spaces(self):
        """공백을 underscore로 변환"""
        result = sanitize_filename("my file name.xlsx")
        assert " " not in result
        assert "_" in result
    
    def test_sanitize_filename_korean(self):
        """한글 문자 유지"""
        result = sanitize_filename("2024년_매출.xlsx")
        assert "2024" in result
        assert "매출" in result
    
    def test_handle_duplicate_filename_no_duplicate(self):
        """파일명이 없을 때"""
        existing = set()
        result = handle_duplicate_filename("file.xlsx", existing)
        assert result == "file.xlsx"
    
    def test_handle_duplicate_filename_with_duplicate(self):
        """파일명 중복 시 suffix 추가"""
        existing = {"file.xlsx"}
        result = handle_duplicate_filename("file.xlsx", existing)
        assert result == "file(1).xlsx"
        
        existing.add(result)
        result = handle_duplicate_filename("file.xlsx", existing)
        assert result == "file(2).xlsx"


# ==================== TEST: API - UPLOAD ====================

class TestUploadAPI:
    """업로드 API 테스트"""
    
    def test_upload_no_file(self, client):
        """파일 없이 요청"""
        response = client.post('/api/upload')
        assert response.status_code == 400
        data = json.loads(response.data)
        assert '파일이' in data['error'] or 'file' in data['error'].lower()
    
    def test_upload_invalid_file_type(self, client):
        """잘못된 파일 형식"""
        data = {
            'file': (io.BytesIO(b'dummy content'), 'test.txt')
        }
        response = client.post('/api/upload', data=data, content_type='multipart/form-data')
        assert response.status_code == 400
        result = json.loads(response.data)
        assert 'XLSX' in result['error'] or 'XLS' in result['error']
    
    def test_upload_valid_2sheets(self, client, sample_excel_2sheets):
        """정상 파일 업로드 (2개 시트)"""
        with open(sample_excel_2sheets, 'rb') as f:
            data = {
                'file': (f, 'sample_2sheets.xlsx')
            }
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')
        
        assert response.status_code == 200
        result = json.loads(response.data)
        
        assert 'session_id' in result
        assert 'temp_file' in result
        assert 'filename' in result
        assert 'sheets' in result
        assert result['sheets'] == ['Sales', 'Expenses']
    
    def test_upload_valid_korean_sheets(self, client, sample_excel_korean):
        """한글 시트명 업로드"""
        with open(sample_excel_korean, 'rb') as f:
            data = {
                'file': (f, 'sample_korean.xlsx')
            }
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')
        
        assert response.status_code == 200
        result = json.loads(response.data)
        assert len(result['sheets']) == 2
        assert '2024년 매출' in result['sheets']


# ==================== TEST: API - SPLIT ====================

class TestSplitAPI:
    """분리 API 테스트"""
    
    def test_split_no_sheets_selected(self, client, sample_excel_2sheets):
        """시트 미선택"""
        # 먼저 업로드
        with open(sample_excel_2sheets, 'rb') as f:
            data = {'file': (f, 'sample_2sheets.xlsx')}
            upload_response = client.post('/api/upload', data=data, content_type='multipart/form-data')
        
        upload_data = json.loads(upload_response.data)
        
        # 시트 미선택하고 분리 요청
        split_data = {
            'session_id': upload_data['session_id'],
            'temp_file': upload_data['temp_file'],
            'filename': upload_data['filename'],
            'sheets': []
        }
        
        response = client.post('/api/split', json=split_data)
        assert response.status_code == 400
        result = json.loads(response.data)
        assert '선택' in result['error']
    
    def test_split_2sheets_success(self, client, sample_excel_2sheets):
        """2개 시트 분리 성공"""
        # 1. 업로드
        with open(sample_excel_2sheets, 'rb') as f:
            data = {'file': (f, 'sample_2sheets.xlsx')}
            upload_response = client.post('/api/upload', data=data, content_type='multipart/form-data')
        
        upload_data = json.loads(upload_response.data)
        
        # 2. 분리 (2개 시트)
        split_data = {
            'session_id': upload_data['session_id'],
            'temp_file': upload_data['temp_file'],
            'filename': upload_data['filename'],
            'sheets': ['Sales', 'Expenses']
        }
        
        response = client.post('/api/split', json=split_data)
        assert response.status_code == 200
        
        # ZIP 파일 반환 확인
        assert response.content_type == 'application/zip'
        assert len(response.data) > 0
    
    def test_split_korean_sheets(self, client, sample_excel_korean):
        """한글 시트명 분리"""
        # 1. 업로드
        with open(sample_excel_korean, 'rb') as f:
            data = {'file': (f, 'sample_korean.xlsx')}
            upload_response = client.post('/api/upload', data=data, content_type='multipart/form-data')
        
        upload_data = json.loads(upload_response.data)
        
        # 2. 분리
        split_data = {
            'session_id': upload_data['session_id'],
            'temp_file': upload_data['temp_file'],
            'filename': upload_data['filename'],
            'sheets': upload_data['sheets']
        }
        
        response = client.post('/api/split', json=split_data)
        assert response.status_code == 200
        assert response.content_type == 'application/zip'


# ==================== TEST: API - HEALTH ====================

class TestHealthAPI:
    """헬스체크 API 테스트"""
    
    def test_health_check(self, client):
        """서버 상태 확인"""
        response = client.get('/api/health')
        assert response.status_code == 200
        data = json.loads(response.data)
        assert data['status'] == 'ok'
        assert 'timestamp' in data


# ==================== MAIN ====================

if __name__ == '__main__':
    pytest.main([__file__, '-v', '--tb=short'])
