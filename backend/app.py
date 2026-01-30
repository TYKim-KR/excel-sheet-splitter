import os
import tempfile
import shutil
import zipfile
import logging
import io
from datetime import datetime, timedelta
from functools import wraps
import uuid

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Alignment, PatternFill, Protection
import signal

# ==================== CONFIG ====================
MAX_FILE_SIZE = 30 * 1024 * 1024  # 30MB
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
UPLOAD_TIMEOUT = 30  # 초
TEMP_CLEANUP_INTERVAL = 3600  # 1시간마다 정리

# ==================== LOGGING ====================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('/tmp/excel_splitter.log') if os.access('/tmp', os.W_OK) else logging.NullHandler()
    ]
)
logger = logging.getLogger(__name__)

# ==================== FLASK APP ====================
app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}})
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# 세션 저장소 (프로덕션: Redis 권장)
SESSION_STORE = {}
CLEANUP_TIME = {}


# ==================== UTILITY FUNCTIONS ====================

def allowed_file(filename):
    """파일 확장자 체크"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def sanitize_filename(filename, max_length=200):
    """
    OS별 파일명 안전화 (Windows, Mac, Linux)
    - 금지 문자 제거
    - 공백을 underscore로 변환
    - 한글/특수문자 이스케이프
    """
    # 파일명과 확장자 분리
    if '.' in filename:
        name, ext = filename.rsplit('.', 1)
    else:
        name = filename
        ext = ''

    # 금지 문자 (Windows: < > : " / \ | ? *)
    forbidden_chars = '<>:"/\\|?*'
    for char in forbidden_chars:
        name = name.replace(char, '_')

    # 공백 → underscore
    name = name.replace(' ', '_')

    # 한글/특수문자 처리: 유니코드 정규화 후 이스케이프 가능 문자만 남김
    # 간단히: 영숫자, underscore, hyphen, 한글 허용
    safe_name = ''
    for char in name:
        if char.isalnum() or char in ('_', '-'):
            safe_name += char
        elif ord(char) >= 0xAC00 and ord(char) <= 0xD7A3:  # 한글 범위
            safe_name += char
        else:
            safe_name += '_'

    # 길이 제한 (확장자 제외)
    if len(safe_name) > max_length:
        safe_name = safe_name[:max_length]

    # 확장자 복원
    if ext:
        return f"{safe_name}.{ext}"
    return safe_name


def handle_duplicate_filename(filename, existing_names):
    """
    파일명 중복 처리: filename(1), filename(2), ...
    """
    if filename not in existing_names:
        return filename

    base, ext = filename.rsplit('.', 1)
    counter = 1

    while True:
        new_name = f"{base}({counter}).{ext}"
        if new_name not in existing_names:
            return new_name
        counter += 1


def cleanup_old_sessions():
    """
    1시간 이상 된 세션 파일 삭제
    """
    now = datetime.now()
    expired = [sid for sid, ts in CLEANUP_TIME.items() if now - ts > timedelta(seconds=TEMP_CLEANUP_INTERVAL)]

    for sid in expired:
        temp_path = SESSION_STORE.get(sid, {}).get('temp_dir')
        if temp_path and os.path.exists(temp_path):
            try:
                shutil.rmtree(temp_path, ignore_errors=True)
                logger.info(f"Cleaned up session: {sid}")
            except Exception as e:
                logger.error(f"Failed to cleanup {sid}: {str(e)}")

        SESSION_STORE.pop(sid, None)
        CLEANUP_TIME.pop(sid, None)


def timeout_handler(signum, frame):
    """타임아웃 처리"""
    raise TimeoutError("처리 시간 초과")


# ==================== API ENDPOINTS ====================

@app.route('/api/health', methods=['GET'])
def health():
    """헬스체크"""
    return jsonify({'status': 'ok', 'timestamp': datetime.now().isoformat()}), 200


@app.route('/api/upload', methods=['POST'])
def upload_file():
    """
    POST /api/upload
    요청: multipart/form-data (file)
    응답: {
        'session_id': str,
        'temp_file': str,
        'filename': str,
        'sheets': [str, ...]
    }
    """
    temp_dir = None

    try:
        # 요청 체크
        if 'file' not in request.files:
            return jsonify({'error': '파일이 없습니다.'}), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({'error': '파일명이 없습니다.'}), 400

        if not allowed_file(file.filename):
            return jsonify({'error': 'XLSX 또는 XLS 파일만 지원합니다.'}), 400

        # 파일 크기 체크
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)

        if file_size > MAX_FILE_SIZE:
            return jsonify({
                'error': f'파일 크기 초과. 최대 {MAX_FILE_SIZE / 1024 / 1024:.0f}MB입니다.'
            }), 400

        # 임시 디렉토리 생성
        temp_dir = tempfile.mkdtemp(prefix='excel_splitter_')
        session_id = os.path.basename(temp_dir)
        temp_file_path = os.path.join(temp_dir, secure_filename(file.filename))

        # 파일 저장
        file.save(temp_file_path)
        logger.info(f"File uploaded: {session_id}, size={file_size} bytes")

        # 워크북 로드 및 시트 목록 추출
        try:
            workbook = openpyxl.load_workbook(temp_file_path, data_only=False)
            sheet_names = workbook.sheetnames
            workbook.close()

            logger.info(f"Sheets extracted: {sheet_names}")

            # 세션 저장
            SESSION_STORE[session_id] = {
                'temp_dir': temp_dir,
                'temp_file': temp_file_path,
                'filename': file.filename,
                'sheets': sheet_names,
                'created_at': datetime.now()
            }
            CLEANUP_TIME[session_id] = datetime.now()

            return jsonify({
                'session_id': session_id,
                'temp_file': temp_file_path,
                'filename': file.filename,
                'sheets': sheet_names
            }), 200

        except openpyxl.utils.exceptions.InvalidFileException:
            shutil.rmtree(temp_dir, ignore_errors=True)
            logger.error(f"Invalid file format: {file.filename}")
            return jsonify({'error': '손상된 엑셀 파일입니다.'}), 400

        except Exception as e:
            shutil.rmtree(temp_dir, ignore_errors=True)
            logger.error(f"Workbook load failed: {str(e)}")
            return jsonify({'error': '파일을 읽을 수 없습니다.'}), 400

    except Exception as e:
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)
        logger.error(f"Upload handler error: {str(e)}")
        return jsonify({'error': '업로드 중 오류가 발생했습니다.'}), 500


@app.route('/api/split', methods=['POST'])
def split_sheets():
    """
    POST /api/split
    요청: {
        'session_id': str,
        'temp_file': str,
        'filename': str,
        'sheets': [str, ...]
    }
    응답: Excel파일 또는 ZIP파일 (다운로드)
    """
    try:
        data = request.get_json()
        session_id = data.get('session_id')
        temp_file = data.get('temp_file')
        filename = data.get('filename')
        selected_sheets = data.get('sheets', [])

        # 유효성 체크
        if not temp_file or not os.path.exists(temp_file):
            return jsonify({'error': '파일을 찾을 수 없습니다.'}), 400

        if not selected_sheets:
            return jsonify({'error': '시트를 1개 이상 선택해주세요.'}), 400

        if len(selected_sheets) > 100:
            return jsonify({'error': '선택 시트가 너무 많습니다. (최대 100개)'}), 400

        logger.info(f"Splitting sheets for session {session_id}: {selected_sheets}")

        # 원본 파일명 (확장자 제외)
        base_filename = os.path.splitext(filename)[0]
        base_filename = sanitize_filename(base_filename)

        # 분리 처리
        output_files = {}
        existing_names = set()

        try:
            source_workbook = openpyxl.load_workbook(temp_file, data_only=False)
        except Exception as e:
            logger.error(f"Failed to load workbook: {str(e)}")
            return jsonify({'error': '파일을 읽을 수 없습니다.'}), 400

        for sheet_name in selected_sheets:
            if sheet_name not in source_workbook.sheetnames:
                logger.warning(f"Sheet not found: {sheet_name}")
                continue

            try:
                # 새 워크북 생성
                new_workbook = openpyxl.Workbook()
                new_sheet = new_workbook.active
                
                # 원본 시트 참조
                source_sheet = source_workbook[sheet_name]

                # 새 시트 제목 (최대 31자)
                new_sheet.title = sheet_name[:31]

                # ===== 데이터 복사 =====
                # 1. 셀 값 및 스타일
                for row in source_sheet.iter_rows():
                    for cell in row:
                        new_cell = new_sheet.cell(row=cell.row, column=cell.column)

                        # 값 복사 (수식 포함)
                        if cell.data_type == 'f':  # 수식
                            new_cell.value = cell.value
                        else:
                            new_cell.value = cell.value

                        # 스타일 복사
                        if cell.has_style:
                            try:
                                new_cell.font = Font(
                                    name=cell.font.name,
                                    size=cell.font.size,
                                    bold=cell.font.bold,
                                    italic=cell.font.italic,
                                    vertAlign=cell.font.vertAlign,
                                    underline=cell.font.underline,
                                    strike=cell.font.strike,
                                    color=cell.font.color
                                )
                            except:
                                pass

                            try:
                                new_cell.border = Border(
                                    left=cell.border.left,
                                    right=cell.border.right,
                                    top=cell.border.top,
                                    bottom=cell.border.bottom,
                                    diagonal=cell.border.diagonal,
                                    diagonal_direction=cell.border.diagonal_direction
                                )
                            except:
                                pass

                            try:
                                new_cell.fill = PatternFill(
                                    fill_type=cell.fill.fill_type,
                                    start_color=cell.fill.start_color,
                                    end_color=cell.fill.end_color,
                                    fgColor=cell.fill.fgColor,
                                    bgColor=cell.fill.bgColor
                                )
                            except:
                                pass

                            try:
                                new_cell.number_format = cell.number_format
                            except:
                                pass

                            try:
                                new_cell.alignment = Alignment(
                                    horizontal=cell.alignment.horizontal,
                                    vertical=cell.alignment.vertical,
                                    text_rotation=cell.alignment.text_rotation,
                                    wrap_text=cell.alignment.wrap_text,
                                    shrink_to_fit=cell.alignment.shrink_to_fit,
                                    indent=cell.alignment.indent
                                )
                            except:
                                pass

                            try:
                                new_cell.protection = Protection(
                                    locked=cell.protection.locked,
                                    hidden=cell.protection.hidden
                                )
                            except:
                                pass

                # 2. 열 너비 복사
                for col_letter in source_sheet.column_dimensions:
                    col_width = source_sheet.column_dimensions[col_letter].width
                    if col_width:
                        new_sheet.column_dimensions[col_letter].width = col_width

                # 3. 행 높이 복사
                for row_num in source_sheet.row_dimensions:
                    row_height = source_sheet.row_dimensions[row_num].height
                    if row_height:
                        new_sheet.row_dimensions[row_num].height = row_height

                # 4. Merged cells 복사
                try:
                    for merged_cell_range in source_sheet.merged_cells.ranges:
                        new_sheet.merge_cells(str(merged_cell_range))
                except:
                    pass

                # 파일명 생성
                safe_sheet_name = sanitize_filename(sheet_name)
                output_filename = f"{base_filename}_{safe_sheet_name}.xlsx"
                output_filename = handle_duplicate_filename(output_filename, existing_names)
                existing_names.add(output_filename)

                # 버퍼에 저장
                output_buffer = io.BytesIO()
                new_workbook.save(output_buffer)
                output_buffer.seek(0)
                output_files[output_filename] = output_buffer.getvalue()

                new_workbook.close()
                logger.info(f"Sheet split completed: {sheet_name} -> {output_filename}")

            except Exception as e:
                logger.error(f"Error splitting sheet '{sheet_name}': {str(e)}")
                continue

        source_workbook.close()

        if not output_files:
            return jsonify({'error': '분리할 수 있는 시트가 없습니다.'}), 400

        # 결과 반환
        if len(output_files) == 1:
            # 파일 1개: 직접 다운로드
            filename, content = list(output_files.items())[0]
            logger.info(f"Single file download: {filename}")

            return send_file(
                io.BytesIO(content),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )

        else:
            # 여러 파일: ZIP 생성
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                for fname, content in output_files.items():
                    zf.writestr(fname, content)

            zip_buffer.seek(0)
            zip_filename = f"{base_filename}_split.zip"
            logger.info(f"ZIP download: {zip_filename} ({len(output_files)} files)")

            return send_file(
                zip_buffer,
                mimetype='application/zip',
                as_attachment=True,
                download_name=zip_filename
            )

    except Exception as e:
        logger.error(f"Split handler error: {str(e)}")
        return jsonify({'error': '처리 중 오류가 발생했습니다.'}), 500

    finally:
        # 오래된 세션 정리
        cleanup_old_sessions()

        # 처리 후 임시 파일 삭제 (선택)
        # 주석 제거 시 다운로드 직후 삭제 (보안 강화)
        # if session_id in SESSION_STORE:
        #     temp_dir = SESSION_STORE[session_id].get('temp_dir')
        #     if temp_dir:
        #         shutil.rmtree(temp_dir, ignore_errors=True)


# ==================== ERROR HANDLERS ====================

@app.errorhandler(413)
def request_entity_too_large(error):
    """파일 크기 초과"""
    return jsonify({
        'error': f'파일이 너무 큽니다. 최대 {MAX_FILE_SIZE / 1024 / 1024:.0f}MB입니다.'
    }), 413


@app.errorhandler(500)
def internal_server_error(error):
    """서버 에러"""
    logger.error(f"Internal error: {str(error)}")
    return jsonify({'error': '서버 오류가 발생했습니다.'}), 500


# ==================== STARTUP ====================

if __name__ == '__main__':
    logger.info("=" * 60)
    logger.info("Excel Sheet Splitter - Backend Server Starting")
    logger.info(f"Max file size: {MAX_FILE_SIZE / 1024 / 1024:.0f}MB")
    logger.info("=" * 60)

    # Flask 실행
    app.run(
        host='0.0.0.0',
        port=5000,
        debug=os.getenv('FLASK_ENV', 'production') == 'development',
        threaded=True
    )
